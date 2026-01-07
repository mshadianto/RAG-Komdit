"""
Document Processor for RAG Komite Audit System
Handles document upload, parsing, and processing
"""
from typing import Dict, Optional, List
import os
import logging
from pathlib import Path
import PyPDF2
from docx import Document
import openpyxl
from backend.database import db
from backend.embeddings import embedding_manager
from config.config import UPLOAD_DIR, PROCESSED_DIR

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DocumentProcessor:
    """Processes various document formats for RAG system"""
    
    SUPPORTED_FORMATS = {
        '.pdf': 'PDF Document',
        '.docx': 'Word Document',
        '.doc': 'Word Document',
        '.txt': 'Text File',
        '.xlsx': 'Excel Spreadsheet',
        '.xls': 'Excel Spreadsheet'
    }
    
    def __init__(self):
        logger.info("Document Processor initialized")
    
    def extract_text_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF file"""
        try:
            text = ""
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            
            logger.info(f"Extracted {len(text)} characters from PDF")
            return text
        except Exception as e:
            logger.error(f"Error extracting text from PDF: {str(e)}")
            raise
    
    def extract_text_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX file"""
        try:
            doc = Document(file_path)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            
            logger.info(f"Extracted {len(text)} characters from DOCX")
            return text
        except Exception as e:
            logger.error(f"Error extracting text from DOCX: {str(e)}")
            raise
    
    def extract_text_from_txt(self, file_path: str) -> str:
        """Extract text from TXT file"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                text = file.read()
            
            logger.info(f"Extracted {len(text)} characters from TXT")
            return text
        except Exception as e:
            logger.error(f"Error extracting text from TXT: {str(e)}")
            raise
    
    def extract_text_from_excel(self, file_path: str) -> str:
        """Extract text from Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text_parts.append(row_text)
            
            text = "\n".join(text_parts)
            logger.info(f"Extracted {len(text)} characters from Excel")
            return text
        except Exception as e:
            logger.error(f"Error extracting text from Excel: {str(e)}")
            raise
    
    def extract_text(self, file_path: str, file_type: str) -> str:
        """Extract text based on file type"""
        file_ext = Path(file_path).suffix.lower()
        
        if file_ext == '.pdf':
            return self.extract_text_from_pdf(file_path)
        elif file_ext in ['.docx', '.doc']:
            return self.extract_text_from_docx(file_path)
        elif file_ext == '.txt':
            return self.extract_text_from_txt(file_path)
        elif file_ext in ['.xlsx', '.xls']:
            return self.extract_text_from_excel(file_path)
        else:
            raise ValueError(f"Unsupported file format: {file_ext}")
    
    def detect_category(self, text: str, filename: str) -> str:
        """Detect document category based on content"""
        text_lower = text.lower()
        filename_lower = filename.lower()
        
        # Define keywords for each category
        category_keywords = {
            "Audit Committee Charter": ["charter", "komite audit", "audit committee", "tata kelola"],
            "Audit Planning": ["perencanaan audit", "audit planning", "risk assessment", "program audit"],
            "Financial Review": ["laporan keuangan", "financial statement", "auditor eksternal", "akuntan publik"],
            "Regulatory": ["peraturan", "regulasi", "ojk", "pasar modal", "psak", "spap"],
            "Banking": ["perbankan", "bank", "bi", "likuiditas", "kredit"],
            "Reporting": ["laporan", "disclosure", "annual report", "pengungkapan"]
        }
        
        # Count keyword matches
        category_scores = {}
        for category, keywords in category_keywords.items():
            score = sum(1 for keyword in keywords if keyword in text_lower or keyword in filename_lower)
            category_scores[category] = score
        
        # Return category with highest score
        if category_scores:
            best_category = max(category_scores.items(), key=lambda x: x[1])
            if best_category[1] > 0:
                return best_category[0]
        
        return "General"
    
    def generate_tags(self, text: str) -> List[str]:
        """Generate tags based on content"""
        tags = []
        text_lower = text.lower()
        
        # Define tag keywords
        tag_keywords = {
            "governance": ["governance", "tata kelola", "pengelolaan"],
            "risk": ["risk", "risiko", "risk management"],
            "compliance": ["compliance", "kepatuhan", "regulasi"],
            "audit": ["audit", "auditor", "pemeriksaan"],
            "financial": ["keuangan", "financial", "laporan keuangan"],
            "internal_control": ["pengendalian intern", "internal control"],
            "ethics": ["etika", "ethics", "kode etik"],
            "transparency": ["transparansi", "transparency", "keterbukaan"]
        }
        
        for tag, keywords in tag_keywords.items():
            if any(keyword in text_lower for keyword in keywords):
                tags.append(tag)
        
        return tags
    
    async def process_document(
        self,
        file_path: str,
        filename: str,
        file_type: str,
        file_size: int
    ) -> Dict:
        """
        Complete document processing pipeline:
        1. Create database entry
        2. Extract text
        3. Detect category and generate tags
        4. Chunk and embed text
        5. Store embeddings
        6. Update document status
        """
        try:
            logger.info(f"Starting to process document: {filename}")
            
            # Step 1: Create document entry in database
            document = await db.create_document(
                filename=filename,
                file_type=file_type,
                file_size=file_size
            )
            document_id = document["id"]
            
            # Step 2: Extract text
            logger.info("Extracting text from document...")
            text = self.extract_text(file_path, file_type)
            
            if not text or len(text.strip()) < 50:
                raise ValueError("Document contains insufficient text content")
            
            # Step 3: Detect category and generate tags
            category = self.detect_category(text, filename)
            tags = self.generate_tags(text)
            
            # Update document with category and tags
            await db.update_document_status(
                document_id=document_id,
                status="processing"
            )
            
            # Update category and tags in database
            from config.config import settings
            supabase_client = db.client
            supabase_client.table(settings.DOCUMENTS_TABLE).update({
                "category": category,
                "tags": tags
            }).eq("id", document_id).execute()
            
            # Step 4: Process document for embedding (chunk and embed)
            logger.info("Generating embeddings...")
            processed_chunks = embedding_manager.process_document_for_embedding(
                text=text,
                document_metadata={
                    "filename": filename,
                    "category": category,
                    "tags": tags
                }
            )
            
            # Step 5: Insert embeddings into database
            logger.info(f"Storing {len(processed_chunks)} chunks in database...")
            await db.insert_embeddings(
                document_id=document_id,
                chunks=processed_chunks
            )
            
            # Step 6: Update document status to processed
            await db.update_document_status(
                document_id=document_id,
                status="processed",
                total_chunks=len(processed_chunks)
            )
            
            logger.info(f"Document processed successfully: {filename}")
            
            return {
                "success": True,
                "document_id": document_id,
                "filename": filename,
                "category": category,
                "tags": tags,
                "total_chunks": len(processed_chunks),
                "text_length": len(text)
            }
            
        except Exception as e:
            logger.error(f"Error processing document: {str(e)}")
            
            # Update document status to error if document was created
            if 'document_id' in locals():
                await db.update_document_status(
                    document_id=document_id,
                    status="error"
                )
            
            return {
                "success": False,
                "error": str(e),
                "filename": filename
            }

# Global document processor instance
document_processor = DocumentProcessor()
